"""Regression coverage for CSV / Excel / JSON / HTML export formats."""

from __future__ import annotations

import csv
import json
import logging
import tempfile
import unittest
from pathlib import Path
from unittest.mock import MagicMock

from database.models import MessageRecord
from services.export_service import ExportService, ExportServiceError


def _synthetic_message(
    *,
    text: str = "hello",
    message_id: int = 1,
    sender_name: str = "alice",
) -> MessageRecord:
    return MessageRecord(
        id=None,
        tg_chat_id=-100123,
        message_id=message_id,
        sender_id=42,
        sender_name=sender_name,
        date="2026-07-22T12:00:00",
        text=text,
        text_preview=text[:80],
        message_type="text",
        has_media=False,
        media_type="",
        file_name="",
        file_size=None,
        is_downloaded=False,
        source_link="https://t.me/c/123/1",
        external_urls="",
        local_path="",
    )


class ExportFormatRegressionTests(unittest.TestCase):
    def setUp(self) -> None:
        self._temp = Path(tempfile.mkdtemp(prefix="tg_export_test_"))
        self._logger = logging.getLogger("test.export")
        self._logger.addHandler(logging.NullHandler())
        self._service = ExportService(
            message_repository=MagicMock(),
            search_service=MagicMock(),
            export_root=self._temp,
            logger=self._logger,
        )

    def tearDown(self) -> None:
        for path in self._temp.glob("*"):
            path.unlink(missing_ok=True)
        self._temp.rmdir()

    def test_empty_results_export_csv_json_headers_only(self) -> None:
        report = self._service.export_message_records(
            [],
            ["csv", "json"],
            base_name="empty_case",
        )
        self.assertEqual(report.total_count, 0)
        self.assertEqual(len(report.files), 2)
        for path_str in report.files:
            path = Path(path_str)
            self.assertTrue(path.exists())
            self.assertTrue(path.is_file())
            if path.suffix == ".csv":
                with path.open(encoding="utf-8-sig", newline="") as handle:
                    rows = list(csv.DictReader(handle))
                self.assertEqual(rows, [])
            elif path.suffix == ".json":
                self.assertEqual(json.loads(path.read_text(encoding="utf-8")), [])

    def test_unicode_text_round_trips_json_csv_html(self) -> None:
        unicode_text = "你好 🎉 café — 测试"
        messages = [
            _synthetic_message(text=unicode_text, message_id=7, sender_name="测试用户"),
        ]
        report = self._service.export_message_records(
            messages,
            ["csv", "json", "html"],
            base_name="unicode_case",
        )
        self.assertEqual(report.total_count, 1)
        for path_str in report.files:
            path = Path(path_str)
            self.assertTrue(path.exists())
            content = path.read_text(encoding="utf-8-sig" if path.suffix == ".csv" else "utf-8")
            self.assertIn(unicode_text, content)
            self.assertIn("测试用户", content)

    def test_excel_export_when_openpyxl_available(self) -> None:
        try:
            import openpyxl  # noqa: F401
        except ImportError:
            self.skipTest("openpyxl not installed")

        messages = [_synthetic_message(text="excel row", message_id=3)]
        report = self._service.export_message_records(
            messages,
            ["xlsx"],
            base_name="excel_case",
        )
        self.assertEqual(report.total_count, 1)
        self.assertEqual(len(report.files), 1)
        path = Path(report.files[0])
        self.assertTrue(path.exists())
        self.assertEqual(path.suffix, ".xlsx")
        from openpyxl import load_workbook

        wb = load_workbook(path)
        sheet = wb.active
        rows = list(sheet.iter_rows(values_only=True))
        self.assertEqual(rows[0][0], "tg_chat_id")
        self.assertIn("excel row", rows[1])

    def test_artifacts_not_left_outside_export_root(self) -> None:
        report = self._service.export_message_records(
            [_synthetic_message()],
            ["json"],
            base_name="local_only",
        )
        for path_str in report.files:
            path = Path(path_str).resolve()
            self.assertTrue(str(path).startswith(str(self._temp.resolve())))

    def test_unknown_format_raises(self) -> None:
        with self.assertRaises(ExportServiceError):
            self._service.export_message_records([], ["pdf"])


if __name__ == "__main__":
    unittest.main()
